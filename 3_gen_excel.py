#!/usr/bin/env python3
# -*- coding: utf-8 -*-


import re
from pathlib import Path
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

INPUT_DIR  = Path("Tool_read_pdf")
PDF_FILE   = Path("O-RAN_E2AP.WG3.TS.E2AP-R004-v07.00.pdf")
OUTPUT_DIR = Path("data_xlsx")
OUTPUT_DIR.mkdir(exist_ok=True)
EXCEL_FILE = OUTPUT_DIR / "data_excel.xlsx"

with open('./msg.config', 'r') as file:
    line1 = file.readline().strip()  # đọc dòng 1
    line2 = file.readline().strip()  # đọc dòng 2 
    PDF_FILE = "./" + line2
    
# ==================== ĐỌC PDF ====================-
# full_text = ""
# with pdfplumber.open(PDF_FILE) as pdf:
#     for page in pdf.pages:
#         text = page.extract_text()
#         if text:
#             full_text += "\n" + text

# full_text = re.sub(r"-\s*\n", "", full_text)
# full_text = re.sub(r"\s+\n", " ", full_text)
# full_text = re.sub(r"\s{2,}", " ", full_text)
# full_text = full_text.replace("…", "...")

# ĐỌC TỪ FILE TXT
with open(PDF_FILE, "r", encoding="utf-8") as f:
    full_text = f.read()
# ==================== DATABASE PRIMITIVE ====================
primitive_db = {}

# 1. INTEGER
for m in re.finditer(r"([A-Za-z0-9\-]+)\s*::=\s*INTEGER\s*\(([^)]+)\)", full_text):
    name = m.group(1).strip()
    content = m.group(2).strip().replace(" ", "")
    if ".." in content:
        before, after = content.split("..", 1)
        min_val = before
        max_val = after.split(",", 1)[0] if "," in after else after
        extra = ",..." if ",..." in m.group(2) else ""
        typ = f"INTEGER ({min_val}..{max_val}{extra})"
        primitive_db[name] = {"type": typ, "min": min_val, "max": max_val, "ext": extra != ""}
    else:
        primitive_db[name] = {"type": f"INTEGER ({content})", "min": "", "max": "", "ext": False}

# 2. ENUMERATED
for m in re.finditer(r"([A-Za-z0-9\-]+)\s*::=\s*ENUMERATED\s*\{([^}]+)\}", full_text):
    name = m.group(1).strip()
    raw = m.group(2)
    clean = re.sub(r"\s+", "", raw)
    has_ext = "..." in clean
    items = [x.strip() for x in clean.replace(",...", "").replace("...", "").split(",") if x.strip()]
    enum_parts = [f'({i},"{item}","{item}")' for i, item in enumerate(items)]
    enum_str = f"[{','.join(enum_parts)}]"
    primitive_db[name] = {
        "type": "ENUMERATED",
        "min": "", "max": "",
        "enum_str": enum_str,
        "ext": has_ext
    }

# 3. OCTET STRING (SIZE)
for m in re.finditer(r"([A-Za-z0-9\-]+)\s*::=\s*(OCTET|BIT)\s+STRING\s*\(SIZE\(([^\)]+)\)\)", full_text, re.I):
    name = m.group(1).strip()
    size_raw = m.group(3).strip()
    size_clean = re.sub(r"\s+", "", size_raw)
    typ = f"{m.group(2).upper()} STRING (SIZE({size_raw}))"
    primitive_db[name] = {"type": typ, "size": size_clean, "ext": False}

# 4. PrintableString - BẮT ĐÚNG CẢ MIN/MAX
for m in re.finditer(r"([A-Za-z0-9\-]+)\s*::=\s*PrintableString\s*\(SIZE\(([^\)]+)\)\)", full_text):
    name = m.group(1).strip()
    size_raw = m.group(2).strip()
    size_clean = re.sub(r"\s+", "", size_raw)
    min_val = max_val = ""
    if ".." in size_clean:
        a, b = size_clean.split("..", 1)
        min_val = a
        max_val = b.split(",", 1)[0] if "," in b else b
    else:
        min_val = max_val = size_clean
    primitive_db[name] = {
        "type": f"PrintableString (SIZE({size_raw}))",
        "min": min_val,
        "max": max_val,
        "ext": True,
        "contain_size": "NULL"  # Theo yêu cầu bạn
    }

# 5. OCTET STRING đơn
for m in re.finditer(r"([A-Za-z0-9\-]+)\s*::=\s*OCTET\s+STRING\b", full_text):
    name = m.group(1).strip()
    if name not in primitive_db:
        primitive_db[name] = {"type": "OCTET STRING", "min": "", "max": "", "ext": False}


# ==================== TẠO EXCEL ====================
wb = Workbook()
ws = wb.active
ws.title = "Primitives"
headers = ["IE_Name", "ASN1_Type", "Min", "Max", "Enum_Items", "Ext.Bit", "Contain_size"]
ws.append(headers)

# Lấy tất cả IE từ bottomup.txt
all_leaves = set()
for f in INPUT_DIR.glob("*bottomup.txt"):
    try:
        with open(f, "r", encoding="utf-8") as f2:
            for line in f2:
                line = line.strip()
                if "." in line and not line.startswith("<"):
                    leaf = line.split(".")[-1].strip()
                    if leaf and not leaf.startswith("----------"):
                        all_leaves.add(leaf)
    except: pass

# Xuất dữ liệu
for leaf in sorted(all_leaves):
    if leaf not in primitive_db:
        continue
    info = primitive_db[leaf]

    asn1_type = info.get("type", "ENUMERATED")
    min_val = info.get("min", "")
    max_val = info.get("max", "")

    # Contain_size
    contain_size = "NULL"
    if "size" in info:
        s = info["size"]
        if ".." in s:
            a, b = s.split("..", 1)
            b = b.split(",", 1)[0] if "," in b else b
            contain_size = f"{a}-{b},..." if ",..." in s else f"{a}-{b}"
        else:
            contain_size = s
    elif info.get("contain_size") is not None:
        contain_size = info["contain_size"]

    enum_items = info.get("enum_str", "")
    ext_bit = "1" if info.get("ext") else "0"

    row = [leaf, asn1_type, min_val, max_val, enum_items, ext_bit, contain_size]
    ws.append(row)

# Format
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=5, max_col=5):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical="top")

for i, col in enumerate(ws.columns, 1):
    max_len = max(len(str(c.value or "")) for c in col)
    ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 80)

wb.save(EXCEL_FILE)

