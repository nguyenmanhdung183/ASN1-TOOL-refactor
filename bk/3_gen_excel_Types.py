#!/usr/bin/env python3
# -*- coding: utf-8 -*-


import re
from pathlib import Path
import pdfplumber
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# ---------- Config ----------
PDF_FILE = Path("O-RAN_E2AP.WG3.TS.E2AP-R004-v07.00.pdf")
BOTTOMUP_DIR = Path("Tool_read_pdf")
OUTPUT_DIR = Path("data_xlsx")
OUTPUT_DIR.mkdir(exist_ok=True)
OUTPUT_XLSX = OUTPUT_DIR / "data_excel.xlsx"
# -------------------------------
# REPLACE VALUES IN ALL SHEETS
# -------------------------------
replace_map = {
    "maxnoofErrors": "256",
    "maxofE2nodeComponents": "1024",
    "maxofRANfunctionID": "256",
    "maxofRICactionID": "16",
    "maxofTNLA": "32",
    "maxofRICrequestID": "1024",
    "maxofRICsubscriptions": "2147483648",
    "maxProtocolIEs": "65536",
}

with open('./msg.config', 'r') as file:
    line1 = file.readline().strip()
    line2 = file.readline().strip()
    PDF_FILE = "./" + line2
    

def replace_values_in_workbook(wb, replace_map):
    """
    replace_map: dict { "giá trị cũ": "giá trị mới" }
    Quét toàn bộ sheet, toàn bộ ô, nếu ô chứa đúng giá trị cũ thì thay bằng giá trị mới.
    """

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if isinstance(val, str):
                    if val in replace_map:
                        cell.value = replace_map[val]

# ---------- Regex helpers ----------
# Capture primitive name plus optional parenthesized constraint (e.g. SIZE(...))
# PRIMITIVE_RE = re.compile(
#     r"\b(BIT\s+STRING|OCTET\s+STRING|INTEGER|ENUMERATED|PrintableString|VisibleString|UTF8String|IA5String|BOOLEAN)"
#     r"(?:\s*\([^\)]*\))?",
#     re.IGNORECASE)

# PRIMITIVE_RE = re.compile(
#     r"\b(BIT\s+STRING|OCTET\s+STRING|INTEGER|ENUMERATED|PrintableString|VisibleString|UTF8String|IA5String|BOOLEAN)"
#     r"(?:\s*\([^\)]*\))?", 
#     re.IGNORECASE
# )
PRIMITIVE_RE = re.compile(
    r"\b(BIT\s+STRING|OCTET\s+STRING|INTEGER|ENUMERATED|PrintableString|VisibleString|UTF8String|IA5String|BOOLEAN)"
    r"(?:\s*\(\s*SIZE\s*\([^\)]*\)\s*\))?",  # Bắt phần SIZE(...) đầy đủ, kể cả lồng ngoặc
    re.IGNORECASE)



BITSTRING_RE = re.compile(r"\bBIT\s+STRING\b", re.IGNORECASE)

FOOTER_PATTERNS = [
    r"©?\s*\d{4}\s+by\s+the\s+O-RAN\s+ALLIANCE",
    r"O-RAN\s+WG\d+",
    r"Your use is subject to the copyright",
    r"Cover page of this specification",
    r"All rights reserved",
]

# ---------- PDF load & clean ----------
def load_pdf_text(path: Path) -> str:
    txt = ""
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                txt += "\n" + t

    txt = re.sub(r"-\s*\n\s*", "", txt)
    txt = txt.replace("\r", "\n")
    txt = re.sub(r"\n{2,}", "\n", txt)
    txt = re.sub(r"[ \t]{2,}", " ", txt)

    lines = []
    for ln in txt.split("\n"):
        s = ln.strip()
        if not any(re.search(pat, s, re.IGNORECASE) for pat in FOOTER_PATTERNS):
            lines.append(ln)

    return "\n".join(lines)


def load_txt_text(path: Path) -> str:
    with open(path, "r", encoding="utf-8") as f:
        txt = f.read()
    return txt



# -------- Balanced brace ----------
def find_balanced_block(text: str, start_brace_idx: int):
    depth = 0
    i = start_brace_idx
    while i < len(text):
        ch = text[i]
        if ch == "{":
            depth += 1
        elif ch == "}":
            depth -= 1
            if depth == 0:
                return i, text[start_brace_idx:i+1]
        i += 1
    return None, ""

# ---------- Multi-line detectors ----------
def detect_sequence_of_singlecontainer_by_lines(text: str):
    blocks = {}
    lines = text.split("\n")
    n = len(lines)
    for i in range(n):
        w = " ".join(lines[i:min(i+3, n)])
        w = re.sub(r"\s+", " ", w).strip()

        m = re.search(
            r"([A-Za-z0-9\-\_]+)\s*::=\s*SEQUENCE"
            r"(?:\s*\(\s*SIZE\s*\([^\)]*\)\s*\))?"
            r"\s*OF\s*(ProtocolIE[-\s]*SingleContainer|SingleContainer)",
            w, re.IGNORECASE
        )
        if m:
            name = m.group(1).strip()

            m_inner = re.search(r"\{\s*\{\s*([A-Za-z0-9\-\_]+)\s*\}\s*\}", w)
            inner = m_inner.group(1).strip() if m_inner else ""

            minv = ""
            maxv = ""
            msize = re.search(r"SIZE\s*\(\s*([0-9]+)(?:\s*\.\.\s*([A-Za-z0-9\-\_]+))?", w)
            if msize:
                minv = msize.group(1)
                if msize.group(2):
                    maxv = msize.group(2)

            blocks[name] = {"kind": "SingleContainer", "inner": inner, "min": minv, "max": maxv}

    return blocks


def detect_container_by_lines(text: str):
    blocks = {}
    lines = text.split("\n")
    n = len(lines)
    for i in range(n):
        w = " ".join(lines[i:min(i+3, n)])
        w = re.sub(r"\s+", " ", w).strip()

        m = re.search(
            r"([A-Za-z0-9\-\_]+)\s*::=\s*(ProtocolIE[-\s]*Container|Container|PairContainer)",
            w, re.IGNORECASE
        )
        if m:
            name = m.group(1).strip()
            m_inner = re.search(r"\{\s*\{\s*([A-Za-z0-9\-\_]+)\s*\}\s*\}", w)
            inner = m_inner.group(1).strip() if m_inner else ""
            blocks[name] = {"kind": "Container", "inner": inner}

    return blocks

# ---------- Find block by name ----------
def find_type_block(name: str, text: str) -> str:
    p_ie = re.compile(
        rf"(^|\n)\s*{re.escape(name)}\s+[A-Za-z0-9\-\_]*\s*E2AP-PROTOCOL-IES\s*::=",
        re.IGNORECASE)
    m_ie = p_ie.search(text)
    if m_ie:
        pos = m_ie.end()
        brace = text.find("{", pos)
        if brace != -1:
            end_idx, blk = find_balanced_block(text, brace)
            if blk:
                return text[m_ie.start(): end_idx+1]
        return text[m_ie.start(): text.find("\n", m_ie.start())]

    p_struct = re.compile(
        rf"(^|\n)\s*{re.escape(name)}\s*::=\s*"
        r"((SEQUENCE\s*(\([^\)]*\))?\s*OF)|SEQUENCE|CHOICE|SET|ProtocolIE[-\s]*SingleContainer|SingleContainer|Container|PairContainer)\b",
        re.IGNORECASE)
    m_struct = p_struct.search(text)
    if m_struct:
        pos = m_struct.end()
        brace = text.find("{", pos)
        if brace != -1:
            end_idx, blk = find_balanced_block(text, brace)
            if blk:
                return text[m_struct.start(): end_idx+1]
        return text[m_struct.start(): text.find("\n", m_struct.start())]

    p_seqof = re.compile(
        rf"(^|\n)\s*{re.escape(name)}\s*::=\s*SEQUENCE\s+OF\s+([A-Za-z0-9\-\_]+)",
        re.IGNORECASE)
    m_seqof = p_seqof.search(text)
    if m_seqof:
        return m_seqof.group(0)

    return ""

# ---------- Split ----------
def split_logical_items(inner: str) -> list:
    items = []
    for ln in inner.split("\n"):
        s = ln.strip()
        if not s:
            continue
        parts = re.split(r",\s*(?=[0-9A-Za-z\-\_])", s)
        for p in parts:
            p2 = p.strip().rstrip(",")
            if p2 and not p2.startswith("..."):
                items.append(p2)
    return items


#======ENUM inline======
def parse_inline_enum_items(type_str):
    """
    Trả về list dạng:
    [(0,"success","success"), (1,"failure","failure"), ...]
    hoặc None nếu không phải ENUM.
    """
    if not isinstance(type_str, str):
        return None

    m = re.search(r"ENUMERATED\s*\{([^}]*)\}", type_str, flags=re.DOTALL | re.IGNORECASE)
    if not m:
        return None

    body = m.group(1)

    raw_items = [
        x.strip().rstrip(",")
        for x in body.split(",")
        if x.strip() not in ("", "...") and not x.strip().startswith("--")
    ]

    out = []
    for idx, it in enumerate(raw_items):
        clean = it.replace("-", "_")
        out.append((idx, clean, clean))

    return out

# ---------- Parse block ----------
def parse_struct_block(name: str, block: str, detected_single_containers, detected_containers):
    rows = []
    if not block:
        return rows
    blk = block
    has_ext = 1 if "..." in blk else 0

    # IE parsing
    if "E2AP-PROTOCOL-IES" in blk:
        inner = blk[blk.find("{")+1: blk.rfind("}")]
        parts = re.findall(r"\{([^{}]+)\}", inner)

        for ent in parts:
            ent = ent.strip()

            mfield = re.search(r"\bID\s+([A-Za-z0-9\-_]+)", ent)
            field_name = mfield.group(1) if mfield else ""

            m_ie = re.search(r"\bTYPE\s+([A-Za-z0-9\-\_]+)", ent)
            ie_type = m_ie.group(1) if m_ie else ""

            mcrit = re.search(r"\bCRITICALITY\s+([A-Za-z0-9\-\_]+)", ent)
            crit = mcrit.group(1) if mcrit else ""

            mpres = re.search(r"\bPRESENCE\s+([A-Za-z0-9\-\_]+)", ent)
            pres = mpres.group(1) if mpres else ""
            optional = "OPTIONAL" if pres.lower().startswith("opt") else ""

            # try parse inline enum if present in the ent text
            enum_items = None
            ein = re.search(r"ENUMERATED\s*\{([^}]*)\}", ent, flags=re.IGNORECASE | re.DOTALL)
            if ein:
                enum_items = parse_inline_enum_items(ein.group(0))

            rows.append({
                "Type_Name": name,
                "ASN1_Type": "IE",
                "Parent_Type": name,
                "Field_Name": field_name,
                "IE_Type": ie_type,
                "Criticality": crit,
                "Optional": optional,
                "Extensible": str(has_ext),
                "Min_Value": "",
                "Max_Value": "",
                "Enum_Item": str(enum_items) if enum_items else ""
            })

        return rows

    # SingleContainer
    if name in detected_single_containers:
        ent = detected_single_containers[name]
        rows.append({
            "Type_Name": name,
            "ASN1_Type": "SingleContainer",
            "Parent_Type": "",
            "Tag/ID": "",
            "Field_Name": name,
            "IE_Type": ent["inner"],
            "Criticality": "",
            "Optional": "",
            "Extensible": str(has_ext),
            "Min_Value": ent["min"],
            "Max_Value": ent["max"],
            "Enum_Item": ""
        })
        return rows

    # Container
    if name in detected_containers:
        ent = detected_containers[name]
        rows.append({
            "Type_Name": name,
            "ASN1_Type": "Container",
            "Parent_Type": "",
            "Tag/ID": "",
            "Field_Name": name,
            "IE_Type": ent["inner"],
            "Criticality": "",
            "Optional": "",
            "Extensible": str(has_ext),
            "Min_Value": "",
            "Max_Value": "",
            "Enum_Item": ""
        })
        return rows

    # SEQUENCE OF generic
    m_seqof = re.search(r"SEQUENCE\s*(\([^\)]*\))?\s*OF\s+([A-Za-z0-9\-\_]+)", blk)
    if m_seqof:
        inner_t = m_seqof.group(2)
        size_token = m_seqof.group(1)

        minv = ""
        maxv = ""
        if size_token:
            msize = re.search(r"SIZE\s*\(\s*([0-9]+)(?:\s*\.\.\s*([A-Za-z0-9\-\_]+))?", size_token)
            if msize:
                minv = msize.group(1)
                if msize.group(2):
                    maxv = msize.group(2)

        rows.append({
            "Type_Name": name,
            "ASN1_Type": f"SEQUENCEOF-{inner_t}",
            "Parent_Type": "",
            "Tag/ID": "",
            "Field_Name": name,
            "IE_Type": "",
            "Criticality": "",
            "Optional": "",
            "Extensible": str(has_ext),
            "Min_Value": minv,
            "Max_Value": maxv,
            "Enum_Item": ""
        })
        return rows

    # STRUCT / CHOICE / SET
    m_top = re.search(r"::=\s*(SEQUENCE|CHOICE|SET)\b", blk)
    if m_top:
        asn1_type = m_top.group(1).upper()

        inner = blk[blk.find("{")+1: blk.rfind("}")]
        items = split_logical_items(inner)

        child_rows = []
        cnt = 0

        for it in items:
            it = it.strip().rstrip(",")

            # TAG form
            m1 = re.match(r"^(\d+)\s+([A-Za-z0-9\-\_]+)\s*(.*)$", it)
            if m1:
                tag = m1.group(1)
                field = m1.group(2)
                rest = m1.group(3).strip()

                # detect inline ENUMERATED { ... } first (this captures braces)
                enum_items = None
                m_inline_enum = re.search(r"(ENUMERATED\s*\{[^}]*\})", rest, flags=re.IGNORECASE | re.DOTALL)
                if m_inline_enum:
                    ie_type = "ENUMERATED"
                    enum_items = parse_inline_enum_items(m_inline_enum.group(1))
                else:
                    # sửa: ưu tiên primitive (kể cả BIT STRING) và lấy full match (kèm SIZE nếu có)
                    m_prim = PRIMITIVE_RE.search(rest)
                    
                    if m_prim:
                        ie_type = m_prim.group(0).strip()
                    else:
                        mtoken = re.search(r"([A-Za-z0-9\-\_]+)", rest)
                        ie_type = mtoken.group(1) if mtoken else field

                optional = "OPTIONAL" if "OPTIONAL" in rest.upper() else ""

                child_rows.append({
                    "Type_Name": name,
                    "ASN1_Type": asn1_type,
                    "Parent_Type": name,
                    "Tag/ID": tag,
                    "Field_Name": field,
                    "IE_Type": ie_type,
                    "Criticality": "",
                    "Optional": optional,
                    "Extensible": "0",
                    "Min_Value": "",
                    "Max_Value": "",
                    "Enum_Item": str(enum_items) if enum_items else ""
                })
                continue

            # simple form
            m2 = re.match(r"^([A-Za-z0-9\-\_]+)\s*(.*)$", it)
            if m2:
                field = m2.group(1)
                rest = m2.group(2).strip()

                enum_items = None
                m_inline_enum = re.search(r"(ENUMERATED\s*\{[^}]*\})", rest, flags=re.IGNORECASE | re.DOTALL)
                if m_inline_enum:
                    ie_type = "ENUMERATED"
                    enum_items = parse_inline_enum_items(m_inline_enum.group(1))
                else:
                    m_prim = PRIMITIVE_RE.search(rest)
                    if m_prim:
                        ie_type = m_prim.group(0).strip()
                    else:
                        mtoken = re.search(r"([A-Za-z0-9\-\_]+)", rest)
                        ie_type = mtoken.group(1) if mtoken else field

                optional = "OPTIONAL" if "OPTIONAL" in rest.upper() else ""

                child_rows.append({
                    "Type_Name": name,
                    "ASN1_Type": asn1_type,
                    "Parent_Type": name,
                    "Tag/ID": "",
                    "Field_Name": field,
                    "IE_Type": ie_type,
                    "Criticality": "",
                    "Optional": optional,
                    "Extensible": "0",
                    "Min_Value": "",
                    "Max_Value": "",
                    "Enum_Item": str(enum_items) if enum_items else ""
                })
                continue

        if asn1_type == "CHOICE":
            for cr in child_rows:
                cr["Tag/ID"] = str(cnt)
                cnt += 1

        for cr in child_rows:
            cr["Extensible"] = str(has_ext)
            rows.append(cr)

        return rows

    return rows

# ---------- MAIN ----------
def main():

    print("Loading & cleaning PDF...")
    pdf_text = load_txt_text(PDF_FILE)

    detected_single = detect_sequence_of_singlecontainer_by_lines(pdf_text)
    detected_cont = detect_container_by_lines(pdf_text)

    leaves = set()
    if not BOTTOMUP_DIR.exists():
        raise FileNotFoundError(f"Bottomup dir missing: {BOTTOMUP_DIR}")

    for f in BOTTOMUP_DIR.glob("*_bottomup.txt"):
        with open(f, "r", encoding="utf-8", errors="ignore") as fh:
            for line in fh:
                line = line.strip()
                if "." not in line:
                    continue
                parts = line.split(".")
                if len(parts) < 2:
                    continue
                leaf = parts[-2].strip()
                if leaf:
                    leaves.add(leaf)

    if OUTPUT_XLSX.exists():
        wb = load_workbook(OUTPUT_XLSX)
    else:
        wb = Workbook()

    headers = ["Type_Name", "ASN1_Type", "Parent_Type", "Tag/ID", "Field_Name",
               "IE_Type", "Criticality", "Optional", "Extensible", "Min_Value", "Max_Value","Enum_Item"]

    if "Types" in wb.sheetnames:
        ws = wb["Types"]

        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)

        if ws.max_row == 0:
            ws.append(headers)

        if ws["A1"].value != headers[0]:
            ws.delete_rows(1, ws.max_row)
            ws.append(headers)
    else:
        ws = wb.create_sheet("Types")
        ws.append(headers)

    # Build TYPES
    for leaf in sorted(leaves):
        block = find_type_block(leaf, pdf_text)
        if not block:
            continue

        rows = parse_struct_block(leaf, block, detected_single, detected_cont)
        if not rows:
            continue

        for r in rows:
            if r.get("ASN1_Type") == "IE" and not r.get("IE_Type"):
                continue
            
            fn = r.get("Field_Name", "").replace("...", "").strip()
            # if IE_Type contains inline ENUM, parse_inline_enum_items already filled Enum_Item in row
            # but also try again from IE_Type string (in case parsing missed earlier)
            enum_items_from_type = parse_inline_enum_items(r.get("IE_Type", "")) 
            enum_cell = r.get("Enum_Item", "")
            if not enum_cell and enum_items_from_type:
                enum_cell = str(enum_items_from_type)

            ws.append([
                r.get("Type_Name", ""),
                r.get("ASN1_Type", ""),
                r.get("Parent_Type", ""),
                r.get("Tag/ID", ""),
                fn,
                r.get("IE_Type", ""),
                r.get("Criticality", ""),
                r.get("Optional", ""),
                r.get("Extensible", "0"),
                r.get("Min_Value", ""),
                r.get("Max_Value", ""),
                enum_cell
            ])

    for col_idx in (5, 6):
        for cell in ws[get_column_letter(col_idx)]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for i, col in enumerate(ws.columns, 1):
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 80)
        


    replace_values_in_workbook(wb, replace_map)


    wb.save(OUTPUT_XLSX)
    print("Saved Types sheet:", OUTPUT_XLSX)

    # =================================================
    # ================  BUILD MESSAGES  ===============
    # =================================================

    def to_camel(x):
        if not x:
            return x
        return x[0].lower() + x[1:]

    ws_types = wb["Types"]
    message_rows = []

    types_data = list(ws_types.iter_rows(min_row=2, values_only=True))

    for row in types_data:
        type_name, asn1_type, parent_type, tag_id, field_name, ie_type,critical,presence, *_ = row

        if asn1_type == "SEQUENCE" and field_name == "protocolIEs":
            msg_base = type_name

            for t_row in types_data:
                t_type_name, t_asn1_type, *_ = t_row[:3]
                if (t_asn1_type == "IE" and
                    (t_type_name.startswith(msg_base) and 
                    (t_type_name.endswith("IEs") or t_type_name.endswith("-IEs"))
                    and not t_type_name.endswith("ItemIEs"))):

                    ie_type = t_row[5]
                    critical = t_row[6]
                    presence = t_row[7]
                    ie_id_const = f"ID_id_{ie_type}"

                    field_camel = t_row[4]

                    if presence != "OPTIONAL":
                        presence = "MANDATORY"
                    
                    message_rows.append([msg_base, ie_id_const, ie_type, field_camel, t_type_name, critical, presence])
            
    msg_headers = ["Message_Name", "IE_ID_Constant", "IE_Type", "Field_Name", "Original_IE_Name", "Critical", "Presence"]

    if "Messages" in wb.sheetnames:
        ws_msg = wb["Messages"]

        if ws_msg.max_row > 1:
            ws_msg.delete_rows(2, ws_msg.max_row - 1)

        if ws_msg.max_row == 0:
            ws_msg.append(msg_headers)

        if ws_msg["A1"].value != msg_headers[0]:
            ws_msg.delete_rows(1, ws_msg.max_row)
            ws_msg.append(msg_headers)
    else:
        ws_msg = wb.create_sheet("Messages")
        ws_msg.append(msg_headers)

    for r in message_rows:
        ws_msg.append(r)

    for i, col in enumerate(ws_msg.columns, 1):
        max_len = max(len(str(c.value or "")) for c in col)
        ws_msg.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 60)

    wb.save(OUTPUT_XLSX)
    print("Saved Messages sheet:", OUTPUT_XLSX)


if __name__ == "__main__":
    main()



